import re
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Knit Concept LTD. (Buyer: SMART BLANKS BD) — Carton বুকিং এক্সেল ফরম্যাট
#   - একটা ফাইলে একাধিক শিট (প্রতি শিট = এক Style/PO), প্রতিটা শিট থেকেই
#     ডাটা নেওয়া হয়।
#   - এক শিটের মধ্যে একাধিক "COLOUR" ব্লক থাকে, প্রতিটা ব্লকের নিজস্ব
#     STYLE/COLOUR/PRIORITY/... হেডার (রং-ভেদে বার বার রিপিট হয়) এবং শেষে
#     একটা 'SUB TOTAL:' রো। পুরো শিটের শেষে 'GRAND TOTAL:' রো, তারপর
#     'TOP/ BOTTOM MEASUREMENT' রো (Top/Bottom সাইজ-গ্রুপ + মেজারমেন্ট)।
#
#   ম্যাপিং (ইউজারের নির্দেশ অনুযায়ী):
#     - Excel STYLE কলাম          -> Template Style No
#     - Sheet Name + "/" + COLOUR -> Template PO ("3502/CHARCOAL HTR"-এর মতো)
#     - COLOUR কলাম               -> Template Reference
#     - SIZE কলাম                 -> Template Pack Type এবং Gmt. Size (দুটোতেই একই ভ্যালু)
#     - LENGTH/WIDTH/HIGHT        -> Template Length/Width/Height (as-is, Inch)
#     - QYANTITY কলাম             -> Template Order Qty
#     - Ply                       -> UI থেকে ম্যানুয়ালি সিলেক্ট (manual_ply)
#     - Item Name                 -> ডিফল্ট 'Master Carton' (Top/Bottom ব্লকের
#       জন্য আলাদা, নিচে দেখুন)
#
#   Raw Data-এ পরে ম্যানুয়ালি চেক করার জন্য প্রতিটা normal লাইনে Excel-এর
#   আসল 'TOP BOTTOM' কলামের ভ্যালুও একটা এক্সট্রা key ('top_bottom_raw')
#   হিসেবে রাখা হচ্ছে (builder.py-র Raw Data শীট line_items-এর সব key
#   অটোমেটিক কলাম হিসেবে দেখায়, তাই এখানে শুধু key যোগ করলেই যথেষ্ট)।
#
#   Top/Bottom অটো-সামারি লাইন:
#     শিটের শেষে 'TOP/ BOTTOM MEASUREMENT' রো-তে জোড়ায় জোড়ায়
#     (label, value) থাকে, যেমন ('S+M', "16''+12''"), ('L+XL', "18''+12''")।
#     প্রতিটা জোড়ার জন্য একটা নতুন লাইন-আইটেম বানানো হয়:
#       - item_name = 'Top Bottom'
#       - style_no / po_no / pack_type / size = 'N/A'
#       - reference = গ্রুপ লেবেল (যেমন 'S+M')
#       - length/width = value থেকে বের করা প্রথম দুইটা সংখ্যা (height নেই)
#       - ply = সবসময় '3' (manual_ply নির্বিশেষে)
#       - qty = ঐ গ্রুপের সাইজগুলোর (যেমন S ও M) 'TOP BOTTOM' কলামের সবগুলো
#         ভ্যালুর যোগফল — পুরো শিটের সব রং/ব্লক মিলিয়ে
# ---------------------------------------------------------------------------


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def _clean(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


def _is_num(v):
    return isinstance(v, (int, float))


def read_knitconcept_style_excel(file_stream, filename='', manual_ply=''):
    """মূল entry point। এই ফরম্যাট না হলে (কোনো শিটেই 'STYLE'/'COLOUR'
    কম্বিনেশন-হেডার না পেলে) খালি লিস্ট [] রিটার্ন করে।"""
    wb = load_workbook(file_stream, data_only=True)
    all_items = []
    ply_value = manual_ply.strip() if manual_ply else 'N/A'

    for sn in wb.sheetnames:
        ws = wb[sn]
        max_row = ws.max_row
        max_col = ws.max_column

        current_style = ''
        current_colour = ''
        # সাইজ-ওয়াইজ (normalize করা key, যেমন 'S', 'M', 'XL') এই শিটের সব
        # রং/ব্লক মিলিয়ে TOP BOTTOM কলামের যোগফল — Top/Bottom গ্রুপ-লাইন
        # বানানোর সময় লাগবে
        size_topbottom_sum = {}
        sheet_had_format = False
        # এই শিটের আইটেম আলাদা টেম্প লিস্টে রাখা হচ্ছে, sheet_had_format
        # শেষ পর্যন্ত False থেকে গেলে (মানে হেডার-সিগনেচার-ই মেলেনি) যাতে
        # কোনো ভুল/মিথ্যা-পজিটিভ রো all_items-এ ঢুকে না যায়
        sheet_items = []

        r = 1
        while r <= max_row:
            c1 = ws.cell(row=r, column=1).value
            c2 = ws.cell(row=r, column=2).value
            n1 = _norm(c1)
            n2 = _norm(c2)

            if n1 == 'style' and n2 == 'colour':
                sheet_had_format = True
                r += 1
                continue

            if n1 in ('subtotal', 'grandtotal'):
                current_style = ''
                current_colour = ''
                r += 1
                continue

            if n1.startswith('topbottommeasurement'):
                # জোড়ায় জোড়ায় (label, value) কলাম 4 থেকে শুরু করে স্ক্যান
                c = 4
                while c + 1 <= max_col:
                    label = _clean(ws.cell(row=r, column=c).value)
                    value = _clean(ws.cell(row=r, column=c + 1).value)
                    if not label and not value:
                        c += 2
                        continue
                    sizes_in_group = [
                        _norm(s) for s in re.split(r'[+/]', label) if _norm(s)
                    ]
                    numbers = re.findall(r'\d+(?:\.\d+)?', value)
                    length = numbers[0] if len(numbers) > 0 else ''
                    width = numbers[1] if len(numbers) > 1 else ''
                    qty_total = sum(size_topbottom_sum.get(sz, 0) for sz in sizes_in_group)
                    if qty_total <= 0:
                        # এই সাইজ-গ্রুপের কোনো ডাটাই এই শিটে নেই (যেমন
                        # 4XL+5XL গ্রুপ ডিফাইন করা আছে কিন্তু এই স্টাইলে
                        # 4XL/5XL কোনো কালারে অর্ডার নেই) — খালি লাইন
                        # যোগ না করে স্কিপ করা হচ্ছে
                        c += 2
                        continue
                    sheet_items.append({
                        'item_name': 'Top Bottom',
                        'ewo_no': 'N/A',
                        'style_no': 'N/A',
                        'po_no': 'N/A',
                        'length': length,
                        'width': width,
                        'height': '',
                        'ply': '3',
                        'qty': qty_total,
                        'pack_type': 'N/A',
                        'reference': label,
                        'size': 'N/A',
                        'remarks': '',
                        'color': '',
                        'delivery_date': '',
                        'measurement_unit': 'Inch',
                        'delivery_place_pdf': '',
                        'delivery_address_pdf': '',
                        'top_bottom_raw': '',
                        '_sheet': sn,
                    })
                    c += 2
                r += 1
                continue

            # ডাটা-রো কিনা চেক (QYANTITY কলাম 9 নিউমেরিক হলেই ডাটা-রো)
            qty_val = ws.cell(row=r, column=9).value
            if _is_num(qty_val):
                style_val = _clean(c1)
                colour_val = _clean(c2)
                if style_val:
                    current_style = style_val
                if colour_val:
                    current_colour = colour_val

                if current_style and current_colour:
                    length = ws.cell(row=r, column=5).value
                    width = ws.cell(row=r, column=6).value
                    height = ws.cell(row=r, column=7).value
                    size_val = _clean(ws.cell(row=r, column=8).value)
                    tb_val = ws.cell(row=r, column=10).value

                    if size_val and _is_num(tb_val):
                        key = _norm(size_val)
                        size_topbottom_sum[key] = size_topbottom_sum.get(key, 0) + tb_val

                    sheet_items.append({
                        'item_name': 'Master Carton',
                        'ewo_no': 'N/A',
                        'style_no': current_style,
                        'po_no': f'{sn}/{current_colour}',
                        'length': _clean(length),
                        'width': _clean(width),
                        'height': _clean(height),
                        'ply': ply_value,
                        'qty': qty_val,
                        'pack_type': size_val,
                        'reference': current_colour,
                        'size': size_val,
                        'remarks': '',
                        'color': '',
                        'delivery_date': '',
                        'measurement_unit': 'Inch',
                        'delivery_place_pdf': '',
                        'delivery_address_pdf': '',
                        'top_bottom_raw': tb_val if tb_val is not None else '',
                        '_sheet': sn,
                    })
            r += 1

        if not sheet_had_format:
            continue  # এই শিট এই ফরম্যাটের না — কোনো আইটেম যোগ হবে না

        all_items.extend(sheet_items)

    return all_items
