import re
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Simba Fashions Limited (Buyer: Macy / MMG) — Carton বুকিং এক্সেল ফরম্যাট
#   - প্রতি ফাইলে সাধারণত এক শিট, এক PO (কিন্তু robustness-এর জন্য সব শিট
#     লুপ করা হয়, যাতে ভবিষ্যতে multi-sheet ফাইল আসলেও কাজ করে)
#   - Item Name সরাসরি কোনো কলামে থাকে না — টেবিলের উপরের ইনফো-ব্লকে কোথাও
#     'Elastic' শব্দ-সহ একটা লাইন থাকলে (যেমন 'Hanger Pack, Need Inside
#     Elastic on Width side') তা থেকে বোঝা যায়:
#       - টেক্সটে 'No ... Elastic' (যেমন 'No Need Inside Elastic') থাকলে
#         -> Master Carton
#       - শুধু 'Elastic'-সহ কোনো লাইন থাকলে (No ছাড়া) -> Elastic Hanger Carton
#       - এরকম কোনো লাইনই না পাওয়া গেলে (দ্বিতীয় sample ফাইলে যেমন নেই)
#         -> ডিফল্ট Master Carton
#   - PO Number: উপরের ইনফো-ব্লকে 'MMG PO# 8868281 NEW YORK'-এর মতো টেক্সট
#     থেকে শুধু নিউমারিক অংশ বের করা হয়
#   - Remarks: উপরের ইনফো-ব্লকে 'Division :' লেবেলের পাশের ভ্যালু (যেমন
#     'MDS', 'HAF') পুরো শিটের জন্য একটাই Remarks হিসেবে বসে
#   - Style No / Pack Type / Reference / Qty / L-W-H — মূল ডাটা-টেবিল থেকে,
#     কিন্তু সাইজ-কলামের সংখ্যা (Size Assortments) PO-ভেদে ভিন্ন হতে পারে
#     (norp-এর মতোই), তাই পরের কলামগুলোর (No of carton booking, Length/Width/Height
#     ইত্যাদি) পজিশন হেডার-লেবেল স্ক্যান করে ডাইনামিকভাবে বের করা হয়
#   - Ply: এই ফরম্যাটে কোনো PLY কলাম নেই, তাই আপলোড ফর্মে ইউজার-এন্টার করা
#     manual_ply ব্যবহার হয় (ঠিক outhouse_extractor.py-এর AEO ফলব্যাকের
#     মতোই — নতুন কোনো লজিক বানানো হয়নি, existing convention অনুসরণ করা হয়েছে)
# ---------------------------------------------------------------------------


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def _clean(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


_PO_RE = re.compile(r'PO\s*#?\s*[:\-]?\s*(\d{4,})', re.I)


def _find_header_row(ws, max_scan=60):
    """'TMCL Carton Label No.' (কলাম A) আর 'PID' (কলাম C) দিয়ে হেডার রো
    খোঁজা হয় — এই কম্বিনেশন norp (PID/COLOR কলাম A/B) বা AEO (PO#/STYLE#)
    ফরম্যাটের সাথে গুলিয়ে যায় না।"""
    for r in range(1, max_scan + 1):
        a = _norm(ws.cell(row=r, column=1).value)
        c = _norm(ws.cell(row=r, column=3).value)
        if a == 'tmclcartonlabelno' and c == 'pid':
            return r
    return None


def _build_dynamic_col_map(ws, header_row, window=4):
    """PID/PPK Code/Description কলাম header_row-এই থাকে, কিন্তু No of carton booking,
    Length/Width/Height ইত্যাদি সাইজ-কলামের সংখ্যার ওপর নির্ভর করে শিফট হয় —
    তাই header_row থেকে কয়েক রো নিচ পর্যন্ত (সব সাব-হেডার সহ) প্রতিটা কলামের
    লেবেল-টেক্সট জোড়া লাগিয়ে স্ক্যান করা হচ্ছে, fixed index ধরে রাখা হচ্ছে না।"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        parts = []
        for rr in range(header_row, header_row + window):
            v = ws.cell(row=rr, column=c).value
            if v is not None and _clean(v):
                parts.append(str(v))
        label = _norm(' '.join(parts))
        if not label:
            continue
        if label == 'pid':
            col_map['style_no'] = c
        elif 'ppkcode' in label:
            col_map['pack_type'] = c
        elif label == 'description':
            col_map['reference'] = c
        elif 'noofcartonbooking' in label:
            col_map['qty'] = c
        elif label.startswith('lengthcm') or label == 'length':
            col_map['length'] = c
        elif label.startswith('widthcm') or label == 'width':
            col_map['width'] = c
        elif label.startswith('heightcm') or label == 'height':
            col_map['height'] = c
    return col_map


def _extract_po_no(ws, header_row):
    """উপরের ইনফো-ব্লকে ('MMG PO# 8868281 NEW YORK'-এর মতো) টেক্সট থেকে
    শুধু নিউমারিক PO নম্বরটা বের করে।"""
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            text = str(v)
            if 'mmg' in text.lower() and 'po' in text.lower():
                m = _PO_RE.search(text)
                if m:
                    return m.group(1)
    return ''


def _extract_remarks(ws, header_row):
    """'Division :' লেবেলের পাশের ভ্যালু (একই রো-তে প্রথম নন-এম্পটি সেল)।"""
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and _norm(v).startswith('division'):
                for c2 in range(c + 1, ws.max_column + 1):
                    v2 = ws.cell(row=r, column=c2).value
                    if v2 is not None and _clean(v2):
                        return _clean(v2)
    return ''


def _classify_item_name(ws, header_row):
    """উপরের ইনফো-ব্লকে 'Elastic' শব্দ-সহ কোনো লাইন খুঁজে Item Name ঠিক করে।
    কিছু না পেলে ডিফল্ট Master Carton।"""
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            t = str(v).lower()
            if 'elastic' not in t:
                continue
            if re.search(r'\bno\b[^.]*elastic', t):
                return 'Master Carton'
            return 'Elastic Hanger Carton'
    return 'Master Carton'


def read_simba_style_excel(file_stream, filename='', manual_ply=''):
    """মূল entry point। এই ফরম্যাট না হলে (হেডার না মিললে) খালি লিস্ট [] রিটার্ন
    করে, যাতে outhouse_extractor.py-এর auto-detect চেইনে পরের ফরম্যাটে
    silently fallback হতে পারে।"""
    wb = load_workbook(file_stream, data_only=True)
    all_items = []
    ply_value = manual_ply.strip() if manual_ply else 'N/A'

    for sn in wb.sheetnames:
        ws = wb[sn]
        header_row = _find_header_row(ws)
        if header_row is None:
            continue  # এই শিট এই ফরম্যাটের না

        col_map = _build_dynamic_col_map(ws, header_row)
        required = ('style_no', 'pack_type', 'reference', 'qty', 'length', 'width', 'height')
        if not all(k in col_map for k in required):
            continue  # প্রত্যাশিত কলাম পাওয়া যায়নি — এই ফরম্যাট না

        po_no = _extract_po_no(ws, header_row)
        remarks = _extract_remarks(ws, header_row)
        item_name = _classify_item_name(ws, header_row)

        qty_col = col_map['qty']
        style_col = col_map['style_no']
        pack_col = col_map['pack_type']
        ref_col = col_map['reference']
        len_col = col_map['length']
        wid_col = col_map['width']
        hgt_col = col_map['height']

        r = header_row + 1
        max_row = ws.max_row
        while r <= max_row:
            qty_val = ws.cell(row=r, column=qty_col).value
            style_val = ws.cell(row=r, column=style_col).value

            # ডাটা-রো চেনার উপায়: qty numeric আর style_no ফাঁকা না —
            # এতে সাব-হেডার রো (From/To ইত্যাদি) আর TOTAL/GRAND TOTAL রো
            # (এগুলোতে style_no কলাম ফাঁকা থাকে) নিজে থেকেই বাদ পড়ে যায়।
            if not isinstance(qty_val, (int, float)) or not _clean(style_val):
                r += 1
                continue

            length = ws.cell(row=r, column=len_col).value
            width = ws.cell(row=r, column=wid_col).value
            height = ws.cell(row=r, column=hgt_col).value
            if length in (None, '') or width in (None, '') or height in (None, ''):
                r += 1
                continue

            all_items.append({
                'item_name': item_name,
                'ewo_no': 'N/A',
                'style_no': _clean(style_val),
                'po_no': po_no,
                'length': _clean(length),
                'width': _clean(width),
                'height': _clean(height),
                'ply': ply_value,
                'qty': qty_val,
                'pack_type': _clean(ws.cell(row=r, column=pack_col).value),
                'reference': _clean(ws.cell(row=r, column=ref_col).value),
                'remarks': remarks,
                'color': '',
                'size': '',
                'delivery_date': '',
                'measurement_unit': 'Cm',
                'delivery_place_pdf': '',
                'delivery_address_pdf': '',
                '_sheet': sn,
            })
            r += 1

    return all_items