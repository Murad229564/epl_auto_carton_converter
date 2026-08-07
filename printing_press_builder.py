import os
from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

from printing_press_extractor import compute_qty_mismatch_warnings

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template_files', 'OrderUploadTemplatePrintingPress.xlsx')

# Gmt. Color, Gmt. Size, Qty মাস্ট-হ্যাভ — Thermal-এর সাথে সামঞ্জস্যপূর্ণ।
# Length/Width/Height/Measurement Unit এখানেও ইচ্ছাকৃতভাবে নেই (থার্মালের
# Measurement-এর মতোই আলাদা confirm-ডায়ালগ দিয়ে হ্যান্ডেল হবে)।
REQUIRED_FIELDS = ['color', 'size', 'qty']


def to_num(v):
    try:
        return float(str(v).replace(',', ''))
    except (ValueError, TypeError):
        return v


def na_if_blank(v):
    v = str(v).strip() if v is not None else ''
    return v if v else 'N/A'


def validate_pp_line_items(line_items):
    """মাস্ট-হ্যাভ ফিল্ড (Gmt. Color/Gmt. Size/Qty) মিসিং থাকলে warning রিটার্ন করে।"""
    warnings = []
    for i, item in enumerate(line_items):
        missing = [f for f in REQUIRED_FIELDS if not str(item.get(f, '')).strip()]
        if missing:
            warnings.append(f"Row {i + 1}: {', '.join(missing)} খালি আছে — চেক করুন")
    return warnings


def _strip_external_links(wb):
    """Thermal/Carton বিল্ডারের মতোই — Excel 'Repairs to...' ওয়ার্নিং এড়াতে
    যেকোনো external link স্থায়ীভাবে মুছে দেওয়া হয়।"""
    try:
        if getattr(wb, '_external_links', None):
            wb._external_links = []
    except Exception:
        pass
    try:
        if hasattr(wb, '_external_references'):
            wb._external_references = []
    except Exception:
        pass


def _write_df_sheet(wb, sheet_name, df):
    if df is None or df.empty:
        return
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_pp_excel(line_items, header_info, out_path,
                    customer_override=None, buyer_override=None,
                    po_override=None, item_name_override=None,
                    delivery_date='', delivery_address='',
                    length='', width='', height=0, measurement_unit='CM',
                    raw_df=None, summary_df=None, warnings=None,
                    remark_place=False, remark_address=False):
    """Printing Press মডিউলের জন্য একটাই .xlsx ফাইল বানায়। শীটগুলো Thermal-এর
    মতোই:
    - Sheet1     : Mapped Template
    - Raw Data   : canonical (মেল্ট করা, প্রতি সাইজ = এক রো) লাইন-আইটেম
    - PO Details : PDF-এ যেভাবে ছিল (wide ফরম্যাট, অরিজিনাল কলাম)
    - PO Summary : PDF-এর page0-এর ST Caow/Rate সামারি টেবিল
    - Warnings   : মাস্ট-হ্যাভ ফিল্ড মিসিং থাকলে

    Item Name: header_info['item_name']-এ PDF থেকে পড়া আসল আইটেম নাম থাকে
    (যেমন 'P.S Tag', 'Poly Sticker') — Thermal-এর মতো ফিক্সড স্ট্রিং না,
    কারণ এই এক্সট্রাক্টরের আন্ডারে ভিন্ন ভিন্ন আইটেম নামে অনেক PO আসবে।
    item_name_override দিয়ে UI থেকে ম্যানুয়ালি ওভাররাইড করা যায় (PDF-এ
    আইটেম নাম না পাওয়া গেলে বা ভুল থাকলে)।

    length/width/height/measurement_unit: থার্মালের measurement ফিল্ডের মতোই
    ম্যানুয়াল ইনপুট — সব রো-তে একই ভ্যালু বসবে। ফাঁকা রাখা হলে (ইউজার
    confirm করার পর) ফাঁকাই থাকবে, warning তৈরি করবে না।
    """
    wb = load_workbook(TEMPLATE_PATH)
    _strip_external_links(wb)
    ws = wb['Sheet1']

    item_name = item_name_override or header_info.get('item_name', '') or 'N/A'

    ws['B2'] = po_override or header_info.get('po_number', '') or 'N/A'
    ws['B3'] = customer_override or header_info.get('customer', '') or 'N/A'
    ws['B4'] = buyer_override or header_info.get('buyer', '') or 'N/A'
    # B5 (Item Group) ও B6 (Business Line) টেমপ্লেটে আগে থেকেই 'Printing Press'
    # ফিক্সড করা আছে — এখানে touch করার দরকার নেই।

    sample_row = 8
    n_cols = 25  # A থেকে Y পর্যন্ত
    style_ref = [copy(ws.cell(row=sample_row, column=c).font) for c in range(1, n_cols + 1)]
    fill_ref = [copy(ws.cell(row=sample_row, column=c).fill) for c in range(1, n_cols + 1)]
    align_ref = [copy(ws.cell(row=sample_row, column=c).alignment) for c in range(1, n_cols + 1)]
    border_ref = [copy(ws.cell(row=sample_row, column=c).border) for c in range(1, n_cols + 1)]
    numfmt_ref = [ws.cell(row=sample_row, column=c).number_format for c in range(1, n_cols + 1)]

    start_row = 8
    for i, r in enumerate(line_items):
        row = start_row + i

        remark_parts = []
        if remark_place and r.get('delivery_place_pdf'):
            remark_parts.append(f"Delivery Place: {r.get('delivery_place_pdf')}")
        if remark_address and r.get('delivery_address_pdf'):
            remark_parts.append(f"Delivery Address: {r.get('delivery_address_pdf')}")
        remarks_val = ' | '.join(remark_parts)

        values = [
            item_name,                                # A Item Name (PDF থেকে ডাইনামিক)
            na_if_blank(r.get('ewo_no')),                # B Gmt. EWO No
            na_if_blank(r.get('style_no')),                # C Gmt. Style No
            na_if_blank(r.get('po_no')),                     # D Gmt. PO
            'All',                                             # E Gmt. Destination
            na_if_blank(r.get('reference')),                    # F Reference/SKU Number (Instruction কলাম থেকে)
            'N/A',                                                # G Pack Type
            na_if_blank(r.get('color')),                           # H Gmt. Color
            na_if_blank(r.get('size')),                              # I Gmt. Size
            length if length != '' else '',                           # J Length (ম্যানুয়াল)
            width if width != '' else '',                               # K Width (ম্যানুয়াল)
            height if height != '' else 0,                                # L Height(If Box) (ম্যানুয়াল)
            measurement_unit or 'CM',                                       # M Measurement In
            'N/A',                                                            # N Paper GSM
            'N/A',                                                              # O Paper Type
            'N/A',                                                                # P Color
            'N/A',                                                                  # Q Type of Print
            'N/A',                                                                    # R No. of Print Color
            'N/A',                                                                      # S No. of Sheet
            r.get('uom') or 'Pcs',                                                        # T UOM
            to_num(r.get('qty', '')),                                                      # U Order Qty
            to_num(r.get('rate', '')) if str(r.get('rate', '')).strip() else '',             # V Rate($)
            delivery_date or '',                                                              # W Delivery Date
            delivery_address or '',                                                             # X Delivery Place
            remarks_val,                                                                          # Y Remarks
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            style_c = min(c, n_cols) - 1
            cell.font = style_ref[style_c]
            cell.fill = fill_ref[style_c]
            cell.alignment = align_ref[style_c]
            cell.border = border_ref[style_c]
            cell.number_format = numfmt_ref[style_c]

    _write_df_sheet(wb, 'Raw Data', pd.DataFrame(line_items))
    _write_df_sheet(wb, 'PO Details', raw_df)
    _write_df_sheet(wb, 'PO Summary', summary_df)

    all_warnings = list(warnings) if warnings else []
    all_warnings += compute_qty_mismatch_warnings(line_items, raw_df, summary_df)
    if all_warnings:
        _write_df_sheet(wb, 'Warnings', pd.DataFrame({'Warning': all_warnings}))

    wb.active = 0
    wb.save(out_path)