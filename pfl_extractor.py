import re
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# PRUDENT FASHION LTD. / Norp Knit Industries Ltd. — Buyer: Kohl's
# Carton বুকিং এক্সেল ফরম্যাট।
#   - একই গ্রুপের দুই কাস্টমারের (নাম আলাদা হতে পারে) একই ফরম্যাট — তাই
#     এই extractor customer name দেখে না, শুধু ফরম্যাট (হেডার সিগনেচার)
#     দেখে চেনে, যেকোনো কাস্টমারের আপলোডেই কাজ করবে।
#   - একাধিক শিট/একাধিক ফাইল থাকতে পারে (Simba-র মতোই), প্রতিটা শিট থেকেই
#     ডাটা নেওয়া হয়।
#   - Item Name টেবিলের বাইরে (উপরে বা নিচে) কোথাও 'ELASTIC' শব্দ-সহ একটা
#     নোট-লাইন থেকে বোঝা যায়:
#       - 'NO NEED ELASTIC' / 'NO ELASTIC' -> Master Carton
#       - 'NEED ELASTIC' / 'WITH ELASTIC' (কিন্তু 'NO' ছাড়া) -> Elastic
#         Hanger Carton
#       - কিছু না পাওয়া গেলে ডিফল্ট Master Carton
#   - PO Number টেবিলের উপরের ইনফো-ব্লকে 'PO NUMBER' লেবেলের পাশে থাকে,
#     এটাই পুরো শিটের জন্য PO No এবং EWO No দুটোতেই বসে (এই ফরম্যাটে আলাদা
#     কোনো EWO নম্বর নেই)।
#   - Style No টেবিলের 'STYLE' কলাম থেকে (উপরের ইনফো-ব্লকের STYLE NUMBER
#     থেকে না — কারণ টেবিলে একাধিক স্টাইল-ভ্যারিয়েন্ট মিশ্রিত থাকতে পারে)।
#   - Reference <- COLOR কলাম। Pack Type <- UPC NUMBER কলাম (এই কলামের
#     লেবেল ফাইল-ভেদে 'ITEM UPC NUMBER' বা 'PACK UPC NUMBER' হতে পারে,
#     দুটোই ধরা হয়)।
#   - Ply সবসময় ফিক্সড ৫ (ইউজারের নির্দেশ অনুযায়ী)।
#   - Measurement দুই রকম লে-আউটে আসতে পারে:
#       (ক) একটা কম্বাইন্ড 'MEASUREMENT' টেক্সট কলাম, যেমন
#           '(L-20) × (W-12) × (H-6.5)' বা '(L)18.5 × (W)13 × (H)5.5 Inc.'
#           — regex দিয়ে পার্স করা হয়।
#       (খ) 'CTN MEASUREMENT INCH' নামের একটা গ্রুপ-হেডার-এর নিচে তিনটা
#           আলাদা numeric কলাম: LENGTH / WIDTH / HEIGHT (মাঝে 'X' লেখা
#           আলাদা কলামে থাকে, ডাটা কলাম না) — এই লে-আউটে সরাসরি সংখ্যা
#           বসানো থাকে, টেক্সট পার্স করা লাগে না।
#     দুটো লে-আউটই সাপোর্ট করা হয় (আগে থেকে কোনটা আসবে জানা যায় না)।
#   - টেমপ্লেটের measurement_unit সবসময় 'Inch' বসবে (CM না)।
#   - STYLE/UPC NUMBER/COLOR/measurement/qty কলামগুলোর অবস্থান (কোন
#     কলাম-ইনডেক্সে) ফাইল-ভেদে ডানে-বামে শিফট হতে পারে (BULK/PREPACK
#     টাইপে STYLE কলাম A-তে শুরু হয়, কিন্তু E-COM (EC) টাইপে B-তে শুরু
#     হয়, কারণ টেবিলের বামে একটা এক্সট্রা ফাঁকা কলাম থাকে) — তাই কোনো
#     fixed column index ধরে রাখা হয় না, সবকিছু header-label স্ক্যান করে
#     ডাইনামিকভাবে বের করা হয়।
# ---------------------------------------------------------------------------


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def _clean(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


# দুই রকম প্যাটার্নই কভার করে: '(L-20) × (W-12) × (H-6.5)' এবং
# '(L)18.5 × (W)13 × (H)5.5 Inc.'
_MEASUREMENT_RE = re.compile(
    r'L[-)]?\s*(\d+(?:\.\d+)?).*?W[-)]?\s*(\d+(?:\.\d+)?).*?H[-)]?\s*(\d+(?:\.\d+)?)',
    re.I | re.S,
)


def _parse_measurement(text):
    if not text:
        return '', '', ''
    m = _MEASUREMENT_RE.search(str(text))
    if m:
        return m.group(1), m.group(2), m.group(3)
    return '', '', ''


def _fmt_num(v):
    if v is None or v == '':
        return ''
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ''
    return str(int(f)) if f == int(f) else str(f)


def _find_header_row(ws, max_scan=60):
    """একই রো-তে পাশাপাশি 'STYLE' -> (কোনো একটা)'...UPC NUMBER' -> 'COLOR'
    এই ক্রমটা খুঁজে বের করে — column-position ধরে রাখা হয় না (BULK/PREPACK-এ
    STYLE কলাম A-তে শুরু হয়, E-COM-এ B-তে), শুধু ক্রমটাই মেলানো হয়। রিটার্ন
    করে (header_row, style_col, upc_col, color_col) অথবা None।"""
    for r in range(1, max_scan + 1):
        n_cols = ws.max_column
        for c in range(1, n_cols + 1):
            if _norm(ws.cell(row=r, column=c).value) != 'style':
                continue
            # STYLE-এর পরে কাছাকাছি একটা '...upcnumber' কলাম খুঁজি
            upc_col = None
            for c2 in range(c + 1, min(c + 4, n_cols + 1)):
                lbl = _norm(ws.cell(row=r, column=c2).value)
                if 'upcnumber' in lbl:
                    upc_col = c2
                    break
            if upc_col is None:
                continue
            # তারপর 'COLOR' কলাম
            color_col = None
            for c3 in range(upc_col + 1, min(upc_col + 3, n_cols + 1)):
                if _norm(ws.cell(row=r, column=c3).value) == 'color':
                    color_col = c3
                    break
            if color_col is None:
                continue
            return r, c, upc_col, color_col
    return None


def _find_measurement_cols(ws, header_row):
    """দুই রকম লে-আউট চেষ্টা করে:
    (ক) header_row-এ exact 'measurement' লেবেল থাকা একটা কম্বাইন্ড টেক্সট কলাম
    (খ) header_row বা তার পরের রো-তে আলাদা LENGTH/WIDTH/HEIGHT কলাম
    রিটার্ন করে dict — {'measurement': col} অথবা {'length':c,'width':c,'height':c}।
    """
    n_cols = ws.max_column
    for c in range(1, n_cols + 1):
        if _norm(ws.cell(row=header_row, column=c).value) == 'measurement':
            return {'measurement': c}

    # split L/W/H — group-header যেকোনো রো-তে (header_row বা header_row+1)
    # থাকতে পারে, তাই দুই রো-ই স্ক্যান করা হচ্ছে।
    for scan_row in (header_row, header_row + 1):
        length_col = width_col = height_col = None
        for c in range(1, n_cols + 1):
            lbl = _norm(ws.cell(row=scan_row, column=c).value)
            if lbl == 'length':
                length_col = c
            elif lbl == 'width':
                width_col = c
            elif lbl == 'height':
                height_col = c
        if length_col and width_col and height_col:
            return {'length': length_col, 'width': width_col, 'height': height_col}
    return {}


def _find_qty_col(ws, header_row):
    n_cols = ws.max_column
    for scan_row in (header_row, header_row + 1):
        for c in range(1, n_cols + 1):
            if 'cartonqty' in _norm(ws.cell(row=scan_row, column=c).value):
                return c
    return None


def _extract_po_no(ws, max_scan=25):
    """'PO NUMBER' লেবেলের একই রো-তে প্রথম numeric ভ্যালু খুঁজে বের করে —
    এই লেবেল আর ভ্যালুর মাঝের কলাম-দূরত্ব ফাইল-ভেদে বদলাতে পারে, তাই fixed
    কলাম না ধরে scan করা হচ্ছে।"""
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and _norm(v) == 'ponumber':
                for c2 in range(c + 1, ws.max_column + 1):
                    v2 = ws.cell(row=r, column=c2).value
                    if isinstance(v2, (int, float)):
                        return str(int(v2)) if float(v2).is_integer() else str(v2)
                    if v2 is not None and _clean(v2) and re.search(r'\d{4,}', str(v2)):
                        m = re.search(r'\d{4,}', str(v2))
                        return m.group(0)
    return ''


def _classify_item_name(ws):
    """পুরো শিটে (টেবিলের উপরে/নিচে) 'ELASTIC' শব্দ-সহ প্রথম লাইনটা খুঁজে
    Item Name ঠিক করে। কিছু না পেলে ডিফল্ট Master Carton।"""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            t = str(v).lower()
            if 'elastic' not in t:
                continue
            if re.search(r'\bno\b[^.]*elastic', t):
                return 'Master Carton'
            return 'Elastic Hanger Carton'
    return 'Master Carton'


def _is_end_marker_row(row_values):
    """টেবিলের শেষ বোঝার মার্কার — পুরনো ফরম্যাটে 'G.TTL', নতুন
    BULK/PREPACK/EC ফরম্যাটে 'Grand Total'।"""
    for v in row_values:
        n = _norm(v)
        if n in ('gttl', 'grandtotal'):
            return True
    return False


def read_pfl_style_excel(file_stream, filename=''):
    """মূল entry point। এই ফরম্যাট না হলে (হেডার না মিললে) খালি লিস্ট [] রিটার্ন
    করে, যাতে outhouse_extractor.py-এর auto-detect চেইনে পরের ফরম্যাটে
    silently fallback হতে পারে।"""
    wb = load_workbook(file_stream, data_only=True)
    all_items = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        found = _find_header_row(ws)
        if found is None:
            continue
        header_row, style_col, upc_col, color_col = found

        qty_col = _find_qty_col(ws, header_row)
        meas_cols = _find_measurement_cols(ws, header_row)
        if qty_col is None or not meas_cols:
            continue  # প্রত্যাশিত কলাম পাওয়া যায়নি — এই ফরম্যাট না

        po_no = _extract_po_no(ws)
        item_name = _classify_item_name(ws)

        # হেডার কয় রো জুড়ে (কিছু ফাইলে একটা রো, কিছুতে গ্রুপ+সাব-লেবেল
        # মিলিয়ে দুই রো) তার ওপর নির্ভর করে ডাটা কোথা থেকে শুরু, সেটা বোঝার
        # সহজ উপায়: header_row-এর ঠিক পরের রো-তে style_col-এ যদি এখনো কোনো
        # লেবেল-জাতীয় টেক্সট (সংখ্যা না) থাকে সেটা সাব-হেডার রো, তাই আরেক
        # রো নিচ থেকে ডাটা শুরু ধরা হবে।
        data_start = header_row + 1
        probe = ws.cell(row=data_start, column=style_col).value
        if probe is not None and not _clean(probe):
            # style_col ফাঁকা কিন্তু আশেপাশে কোথাও লেবেল-টেক্সট আছে এমন
            # হলে এটা সাব-হেডার রো — নিচের দিকে এক রো সরিয়ে দেখা হচ্ছে
            pass  # নিচের while-loop-এর data-row detection নিজেই এটা হ্যান্ডেল করবে

        r = data_start
        max_row = ws.max_row
        while r <= max_row:
            row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if _is_end_marker_row(row_values):
                break  # টেবিলের শেষ — এর পরের সব রো ফুটার/নোট, ডাটা না

            qty_val = ws.cell(row=r, column=qty_col).value
            style_val = ws.cell(row=r, column=style_col).value

            # data-row চেনার উপায়: qty numeric আর style_no ফাঁকা না —
            # এতে Grand Total/G.TTL রো আর সাব-হেডার/নোট-লাইনগুলো (যেখানে
            # style_no কলাম ফাঁকা থাকে) নিজে থেকেই বাদ পড়ে যায়।
            if not isinstance(qty_val, (int, float)) or not _clean(style_val):
                r += 1
                continue

            if 'measurement' in meas_cols:
                length, width, height = _parse_measurement(
                    ws.cell(row=r, column=meas_cols['measurement']).value
                )
            else:
                length = _fmt_num(ws.cell(row=r, column=meas_cols['length']).value)
                width = _fmt_num(ws.cell(row=r, column=meas_cols['width']).value)
                height = _fmt_num(ws.cell(row=r, column=meas_cols['height']).value)

            if not length:
                r += 1
                continue

            all_items.append({
                'item_name': item_name,
                'ewo_no': po_no,
                'style_no': _clean(style_val),
                'po_no': po_no,
                'length': length,
                'width': width,
                'height': height,
                'ply': '5',  # ইউজারের নির্দেশ অনুযায়ী — ফিক্সড ৫ প্লাই
                'qty': qty_val,
                'pack_type': _clean(ws.cell(row=r, column=upc_col).value),
                'reference': _clean(ws.cell(row=r, column=color_col).value),
                'remarks': '',
                'color': '',
                'size': '',
                'delivery_date': '',
                'measurement_unit': 'Inch',
                'delivery_place_pdf': '',
                'delivery_address_pdf': '',
                '_sheet': sn,
            })
            r += 1

    return all_items