import re
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Norp Knit-স্টাইল আউট হাউজ Carton বুকিং এক্সেল — এই ফরম্যাটটা AEO-স্টাইল
# ফ্ল্যাট টেবিলের থেকে সম্পূর্ণ আলাদা:
#   - একই ফাইলে একাধিক শিট থাকতে পারে (প্রতিটা শিট থেকেই ডাটা নিতে হবে)
#   - প্রতিটা শিটে PID/COLOR/MMG PO/PCK অনুযায়ী ব্লক-বেসড ডাটা (এই কলামগুলো
#     শুধু প্রতি ব্লকের প্রথম রো-তে থাকে, forward-fill দরকার)
#   - সাইজ-কলামের সংখ্যা শিট-ভেদে ভিন্ন হয় (৬টা থেকে ৯টা পর্যন্ত দেখা গেছে),
#     তাই তার পরের 'Carton qty'/'Measurment'/'Remarks' কলামের পজিশনও শিট-ভেদে
#     শিফট হয়ে যায় — তাই এগুলো হেডার-লেবেল স্ক্যান করে ডাইনামিকভাবে বের করা হয়
#     (fixed column index ধরে রাখা হয় না)।
#   - Item Name PDF-এর মতো সরাসরি কোনো কলামে থাকে না, বরং 'Remarks' কলামের
#     ('Need Elastic...' / 'No Need Elastic...') টেক্সট থেকে বুঝে নিতে হয়।
# ---------------------------------------------------------------------------


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def _clean(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


_MEASUREMENT_RE = re.compile(
    r'L\s*[-:]?\s*(\d+(?:\.\d+)?)\s*CM?\s*[×xX*]\s*W\s*[-:]?\s*(\d+(?:\.\d+)?)\s*CM?\s*'
    r'[×xX*]\s*H\s*[-:]?\s*(\d+(?:\.\d+)?)\s*CM?',
    re.I,
)
_MEASUREMENT_RE_PLAIN = re.compile(
    r'(\d+(?:\.\d+)?)\s*CM?\s*[×xX*]\s*(\d+(?:\.\d+)?)\s*CM?\s*[×xX*]\s*(\d+(?:\.\d+)?)\s*CM?',
    re.I,
)


def _parse_measurement(text):
    """'L-58 CM X W-50 CM X H-12 CM' বা '32CM X 29CM X 20CM' -> (58,50,12) —
    এই ফরম্যাটে মেজারমেন্ট সবসময় CM-এই থাকে (ইউজার নিশ্চিত করেছেন)।"""
    if not text:
        return '', '', ''
    m = _MEASUREMENT_RE.search(str(text))
    if m:
        return m.group(1), m.group(2), m.group(3)
    m = _MEASUREMENT_RE_PLAIN.search(str(text))
    if m:
        return m.group(1), m.group(2), m.group(3)
    return '', '', ''


def _classify_item_name(remark_text):
    """Remarks কলামের টেক্সট থেকে Item Name বুঝে নেয় (ইউজারের নির্দেশ
    অনুযায়ী) — 'No Need Elastic' পেলে Master Carton, শুধু 'Need Elastic'
    (কিন্তু 'No' ছাড়া) পেলে Elastic Hanger Carton। None রিটার্ন করলে বোঝা
    যায় classify করার মতো কিছু নেই এই রো-তে (আগের ভ্যালুই বহাল থাকবে)।"""
    t = str(remark_text or '').lower()
    if 'no need elastic' in t:
        return 'Master Carton'
    if 'need elastic' in t:
        return 'Elastic Hanger Carton'
    return None


def _find_header_row(ws, max_scan=50):
    """'PID' (কলাম A) আর 'COLOR' (কলাম B) দিয়ে হেডার রো খোঁজা হয় — row
    position ধরে রাখা হয় না, যাতে ভবিষ্যতে অন্য বুকিং শিটে হেডার একটু
    উপরে-নিচে হলেও কাজ করে।"""
    for r in range(1, max_scan + 1):
        a = _norm(ws.cell(row=r, column=1).value)
        b = _norm(ws.cell(row=r, column=2).value)
        if a == 'pid' and b.startswith('color'):
            return r
    return None


def _build_dynamic_col_map(ws, header_row):
    """Carton qty/Measurement/Remarks কলামের পজিশন শিট-ভেদে ভিন্ন (সাইজ
    কলামের সংখ্যার ওপর নির্ভর করে) — তাই হেডার রো-র লেবেল টেক্সট স্ক্যান
    করে ডাইনামিকভাবে বের করা হচ্ছে, fixed index ধরে রাখা হচ্ছে না।"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        label = _norm(ws.cell(row=header_row, column=c).value)
        if not label:
            continue
        if 'cartonqty' in label:
            col_map['carton_qty'] = c
        elif label.startswith('measur'):
            col_map['measurement'] = c
        elif label == 'remarks':
            col_map['remarks'] = c
        elif 'deliveryplace' in label:
            col_map['delivery_place'] = c
    return col_map


def read_norp_style_sheet(ws, sheet_name=''):
    """একটা শিট থেকে canonical লাইন-আইটেম বের করে। হেডার না পাওয়া গেলে
    (এই ফরম্যাটের শিট না হলে) খালি লিস্ট রিটার্ন করে — এতে ডাকা যায়
    'এই ফরম্যাট কিনা' যাচাই করার জন্যও।"""
    header_row = _find_header_row(ws)
    if header_row is None:
        return []

    col_map = _build_dynamic_col_map(ws, header_row)
    if 'carton_qty' not in col_map or 'measurement' not in col_map:
        return []  # প্রত্যাশিত কলাম পাওয়া যায়নি — এই ফরম্যাট না

    qty_col = col_map['carton_qty']
    meas_col = col_map['measurement']
    remarks_col = col_map.get('remarks')

    items = []
    current_pid = ''
    current_color = ''
    current_po = ''
    current_pck = ''
    current_item_name = 'Master Carton'  # ডিফল্ট — প্রথম remark না পাওয়া পর্যন্ত

    r = header_row + 2  # হেডার (২ রো: মূল লেবেল + সাইজ-সাব-লেবেল)-এর পরই ডাটা শুরু
    max_row = ws.max_row
    while r <= max_row:
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        qty = ws.cell(row=r, column=qty_col).value
        measurement = ws.cell(row=r, column=meas_col).value
        remark = ws.cell(row=r, column=remarks_col).value if remarks_col else None

        a_str = _clean(a)
        if a_str.upper() == 'TOTAL':
            r += 1
            continue
        if all(v is None for v in (a, b, c, d, e, qty)):
            r += 1
            continue

        if a_str:
            current_pid = a_str
        if b is not None and _clean(b):
            current_color = _clean(b)
        if c is not None and _clean(c):
            current_po = _clean(c)
        if d is not None and _clean(d):
            current_pck = _clean(d)
        if remark is not None and _clean(remark):
            classified = _classify_item_name(remark)
            if classified:
                current_item_name = classified

        e_str = _clean(e)
        if not e_str or qty is None:
            r += 1
            continue

        length, width, height = _parse_measurement(measurement)
        if not length:
            r += 1
            continue

        items.append({
            'item_name': current_item_name,
            'ewo_no': 'N/A',
            'style_no': current_pid,
            'po_no': current_po,
            'length': length,
            'width': width,
            'height': height,
            'ply': '5',  # ইউজারের নির্দেশ অনুযায়ী — এই কাস্টমারের জন্য সবসময় ফিক্সড ৫ প্লাই
            'qty': qty,
            'pack_type': e_str,
            'reference': current_color,
            'remarks': current_pck,
            'color': '',
            'size': '',
            'delivery_date': '',
            'measurement_unit': 'Cm',
            'delivery_place_pdf': '',
            'delivery_address_pdf': '',
            '_sheet': sheet_name,
        })
        r += 1

    return items


def read_norp_style_excel(file_stream, filename=''):
    """ফাইলের প্রতিটা শিট থেকে ডাটা বের করে একটাই লিস্টে মিলিয়ে দেয়
    (ইউজারের নির্দেশ: 'প্রত্যেকটা সিট থেকে ডাটা নিয়ে টেমপ্লেটে বসানো')।
    কোনো শিট এই ফরম্যাটের না হলে (হেডার না মিললে) সেটা চুপচাপ স্কিপ হয়ে যায়।"""
    wb = load_workbook(file_stream, data_only=True)
    all_items = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        items = read_norp_style_sheet(ws, sn)
        all_items.extend(items)
    return all_items
